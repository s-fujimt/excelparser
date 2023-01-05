import openpyxl
import json
from libs.color_helper import theme_and_tint_to_rgb
import zipfile
import xml.etree.ElementTree as ET
import sys
from datetime import datetime
from openpyxl.xml.functions import fromstring, QName

class ExcelParser:
    excel_path = None
    workbook = None
    custom_index  = None
    current_sheet = None
    current_sheet_number = 0
    current_sheet_ranges = None
    empty_rows = 0
    empty_columns = 0

    def __open_workbook(self, excel_path):
        try:
            self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        except:
            print("Error opening workbook", sys.exc_info()[0])
            return None

    def __set_custom_index(self, color):
        try:
            self.custom_index = [rgb.attrib['rgb'] for rgb in color]
        except:
            print("Error setting custom index", sys.exc_info()[0])
            return None

    def __check_for_custom_index(self, filepath):
        try:
            with zipfile.ZipFile(filepath) as zgood:
                styles_xml = zgood.read('xl/styles.xml')
                root = ET.fromstring(styles_xml)
                [[self.__set_custom_index(color) for color in child if 'indexedColors' in color.tag] for child in root if 'colors' in child.tag]
        except:
            print("Error checking for custom index", sys.exc_info()[0])
            return None

    def __get_default_font_data(self):
        try:
            return {
                "font": self.current_sheet.cell(row=1, column=1).font.name,
                "size": int(self.current_sheet.cell(row=1, column=1).font.size)
            }
        except:
            print("Error getting default font data", sys.exc_info()[0])
            return {}

    def __is_merged_cell(self, cell, merged_ranges):
        for merged_range in merged_ranges:
            if cell.coordinate in merged_range:
                self.current_range = {
                    "columns": merged_range.size["columns"],
                    "rows": merged_range.size["rows"]
                }
                return True
        return False

    def __get_first_cells_of_merged_ranges(self):
        return [merged_range.start_cell.coordinate for merged_range in self.current_sheet_ranges]

    def __get_merged_cell_data(self):
        try:
            cell_data = {}
            if self.current_range["columns"] > 1:
                cell_data["colspan"] = self.current_range["columns"]
            if self.current_range["rows"] > 1:
                cell_data["rowspan"] = self.current_range["rows"]
            return cell_data
        except:
            print("Error getting merged cell data", sys.exc_info()[0])
            return {}

    def __get_cell_alignment(self, cell):
        try:
            alignment = {}
            if cell.alignment:
                if cell.alignment.horizontal in ["center", "left", "right"]:
                    alignment["horizontal"] = cell.alignment.horizontal
                if cell.alignment.vertical in ["center", "bottom", "top"]:
                    alignment["vertical"] = cell.alignment.vertical
            return alignment
        except:
            print("Error getting cell alignment", sys.exc_info()[0])
            return {}
    
    def __get_color_from_theme(self, color_data):
        try:
            color = theme_and_tint_to_rgb(self.workbook, color_data.theme, color_data.tint)
            return color
        except:
            print("Error getting color from theme")
            return {}

    def __get_color_data(self, color_data):
        try:
            color = None
            if color_data.type == "rgb":
                if color_data.rgb == "00000000":
                    color = "FFFFFF"
                else:
                    color = color_data.rgb[2:]
            if color_data.type == "indexed":
                if color_data.indexed == 63:
                    color = "FFFFFF"
                if color_data.indexed == 64:
                    color = "000000"
                else:
                    if self.custom_index:
                        Colors = self.custom_index
                    else:
                        Colors = openpyxl.styles.colors.COLOR_INDEX
                    color = Colors[color_data.indexed][2:]
            if color_data.type == "theme":
                color = self.__get_color_from_theme(color_data)
            return f"#{color}"
        except:
            print("Error getting color data (" + str(color_data) + ")") 
            return None

    def __get_cell_font_data(self, cell):
        try:
            cell_font_data = {}
            default_font_data = self.__get_default_font_data()
            if cell.font:
                if cell.font.name != default_font_data["font"]:
                    cell_font_data["font"] = cell.font.name
                if cell.font.size and cell.font.size != default_font_data["size"]:
                    cell_font_data["size"] = int(cell.font.size)
                if cell.font.bold:
                    cell_font_data["style"] = "bold"
                if cell.font.underline:
                    cell_font_data["underline"] = cell.font.underline
                if cell.font.strikethrough:
                    cell_font_data["strikethrough"] = cell.font.strikethrough
                if cell.font.color:
                    color = self.__get_color_data(cell.font.color)
                    if color and color != "#000000":
                        cell_font_data["color"] = color
            return cell_font_data
        except:
            print("Error getting cell font data")
            return {}

    def __get_border_style(self, border_style):
        if border_style == "medium":
            return "thick"
        if border_style == "thick":
            return "extrathick"
        if border_style == "double":
            return "double"
        else:
          return "single"

    def __set_border(self, cell, direction_list):
        try:
            border_top, border_right, border_bottom, border_left = [], [], [], []
            border = cell.border
            direction = direction_list[0]
            partner = direction_list[1]
            neighbor = None

            if getattr(border, direction) and getattr(border, direction).style and getattr(border, direction).color:
                locals()[f"border_{direction}"].append(self.__get_border_style(getattr(border, direction).style))
                locals()[f"border_{direction}"].append(self.__get_color_data(getattr(border, direction).color))
            else:
                if cell.row == 1 and direction == "top":
                    return False
                if cell.column == 1 and direction == "left":
                    return False
                if direction == "top":
                    neighbor = self.current_sheet.cell(row=cell.row-1, column=cell.column)
                elif direction == "right":
                    neighbor = self.current_sheet.cell(row=cell.row, column=cell.column+1)
                elif direction == "bottom":
                    neighbor = self.current_sheet.cell(row=cell.row+1, column=cell.column)
                elif direction == "left":
                    neighbor = self.current_sheet.cell(row=cell.row, column=cell.column-1)
                if getattr(neighbor.border, partner) and getattr(neighbor.border, partner).style and getattr(neighbor.border, partner).color:
                    locals()[f"border_{direction}"].append(self.__get_border_style(getattr(neighbor.border, partner).style))
                    locals()[f"border_{direction}"].append(self.__get_color_data(getattr(neighbor.border, partner).color))
            return locals()[f"border_{direction}"]
        except:
            print("Error setting border", sys.exc_info()[0])
            return []

    # def __set_merged_border(self, cell, direction):
    #     current_range = list(filter(lambda x: cell.coordinate in x, self.current_sheet.merged_cells.ranges))
    #     end_cell = self.current_sheet.cell(row=current_range[0].max_row, column=current_range[0].max_col) if current_range else None

    #     xmlns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    #     with zipfile.ZipFile(self.excel_path) as zgood:
    #         sheet_xml = zgood.read('xl/worksheets/sheet{0}.xml'.format(self.current_sheet_number)) 
    #         root = ET.fromstring(sheet_xml)
    #         sheet_data = root.find(QName(xmlns, 'sheetData').text)
    #         wanted_style = None
    #         if sheet_data is not None:
    #             rows = sheet_data.findall(QName(xmlns, 'row').text)
    #             for r in rows:
    #                 if r.get('r') == str(end_cell.row):
    #                     for c in r:
    #                         if c.get('r') == end_cell.coordinate:
    #                             wanted_style = c.get("s")
    #                             break
    #                 if wanted_style:
    #                     break
    #             styles_xml = zgood.read('xl/styles.xml')
    #             root = ET.fromstring(styles_xml)
    #             cellxfs = root.find(QName(xmlns, 'cellXfs').text)
    #             if cellxfs is not None:
    #                 xfNodes = cellxfs.findall(QName(xmlns, 'xf').text)
    #                 xfNode = xfNodes[int(wanted_style)]
    #                 borderId = xfNode.get("borderId")
    #                 if borderId:
    #                     borders = root.find(QName(xmlns, 'borders').text)
    #                     if borders is not None:
    #                         borderNodes = borders.findall(QName(xmlns, 'border').text)
    #                         borderNode = borderNodes[int(borderId)]
    #                         dir_node = borderNode.find(QName(xmlns, direction).text)
    #                         if dir_node is not None:
    #                             locals()[f"border_{direction}_style"] = self.__get_border_style(dir_node.get("style"))

    #                             col_node = dir_node.find(QName(xmlns, 'color').text)
    #                             if col_node is not None:
    #                                 color = None
    #                                 if col_node.get("rgb"):
    #                                     rgb = col_node.get("rgb")
    #                                     type = "rgb"
    #                                     color = openpyxl.styles.colors.Color(rgb=rgb, type=type)
    #                                 elif col_node.get("indexed"):
    #                                     indexed = col_node.get("indexed")
    #                                     type = "indexed"
    #                                     color = openpyxl.styles.colors.Color(indexed=indexed, type=type)
    #                                 elif col_node.get("auto"):
    #                                     auto = col_node.get("auto")
    #                                     type = "auto"
    #                                     color = openpyxl.styles.colors.Color(auto=auto, type=type)
    #                                 elif col_node.get("theme"):
    #                                     theme = col_node.get("theme")
    #                                     tint = col_node.get("tint")
    #                                     type = "theme"
    #                                     color = openpyxl.styles.colors.Color(theme=theme, tint=tint, type=type)
    #                                 locals()[f"border_{direction}_color"] = self.__get_color_data(color)

    #                                 locals()[f"border_{direction}"] = [locals()[f"border_{direction}_style"], locals()[f"border_{direction}_color"]]

    #                                 return locals()[f"border_{direction}"]

    def __get_cell_border_data(self, cell, is_merged_cell):
        try:
            cell_border_data, outline = {}, {}

            border = cell.border

            if is_merged_cell:
                border_top = self.__set_border(cell, ("top", "bottom"))
                border_left = self.__set_border(cell, ("left", "right"))
                border_bottom = self.__set_border(cell, ("bottom", "top"))
                border_right = self.__set_border(cell, ("right", "left"))
                # border_bottom = self.__set_merged_border(cell, "bottom")
                # border_right = self.__set_merged_border(cell, "right")
                                                
            else:
                directions = [("top", "bottom"), ("right", "left"), ("bottom", "top"), ("left", "right")]
                [border_top, border_right, border_bottom, border_left] = [self.__set_border(cell, direction) for direction in directions]

            if border_top == border_right == border_bottom == border_left and len(border_top) > 0:
                border = border_top
                if len(border) == 1:
                    cell_border_data["outline"] = {"style": border[0]}
                if len(border) == 2:
                    cell_border_data["outline"] = {"style": border[0], "color": border[1]}
            else:
                for direction in ["top", "right", "bottom", "left"]:
                    if locals()[f"border_{direction}"]:
                        if len(locals()[f"border_{direction}"]) == 1:
                            cell_border_data[direction] = {"style": locals()[f"border_{direction}"][0]}
                        if len(locals()[f"border_{direction}"]) == 2:
                            cell_border_data[direction] = {"style": locals()[f"border_{direction}"][0], "color": locals()[f"border_{direction}"][1]}

            if outline:
                cell_border_data["outline"] = outline

            return cell_border_data
        except:
            print("Error getting cell border data (" + cell.coordinate + ")" + sys.exc_info()[0] + " " + str(sys.exc_info()[1]))
            return {}

    def __get_fill_color(self, cell):
        try:
            if cell.fill:
                if cell.fill.start_color:
                    color = self.__get_color_data(cell.fill.start_color)
                    if color and color != "#FFFFFF":
                        return color
            return None
        except:
            print("Error getting cell fill color (" + str(cell.coordinate) + ")")
            return None

    def __map_cell_data(self, cell):
        try:
            cell_data = {
                "colnumber": cell.coordinate[0]
            }
            
            value = cell.value
            if value != None:
                cell_data["value"] = cell.value.strftime("%Y/%m/%d") if cell.is_date else value
                if cell._style[3] in [50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60]:
                    dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + cell_data["value"] - 2).strftime("%Y/%m/%d")
                    cell_data["value"] = dt

            is_merged_cell = self.__is_merged_cell(cell, self.current_sheet_ranges)
            if is_merged_cell:
                if cell.coordinate not in self.__get_first_cells_of_merged_ranges():
                    return {}
                cell_data.update(self.__get_merged_cell_data())

            alignment = self.__get_cell_alignment(cell)
            if alignment:
                cell_data["alignment"] = alignment
            
            font = self.__get_cell_font_data(cell)
            if font:
                cell_data["font"] = font if font else None

            border = self.__get_cell_border_data(cell, is_merged_cell)
            if border:
                cell_data["border"] = border

            fill_color = self.__get_fill_color(cell)
            if fill_color:
                cell_data["fill"] = {"color":fill_color}

            if not (cell_data.get("value") or cell_data.get("border") or cell_data.get("fill")):
                return {}

            return cell_data
        except:
            print("Error getting cell data (" + cell.coordinate + ")")
            print(sys.exc_info()[0])
            return {}

    def __cell_data_wrapper(self, cell):
        cell_data = self.__map_cell_data(cell)
        self.empty_columns = self.empty_columns + 1 if cell_data == {} else 0
        return cell_data if cell_data else None

    def __get_row_data(self, row):
        self.empty_columns = 0
        try:
            row_data = {
                "linenumber": row[0].row
            }

            columns = [self.__cell_data_wrapper(cell) for cell in row if self.empty_columns < 50 and cell.column < self.current_sheet.max_column+2]
            columns = [column for column in columns if column]
            if columns:
                row_data["columns"] = columns

            return row_data
        except:
            print("Error getting row data")
            return {}

    def __row_data_wrapper(self, row):
        row_data = self.__get_row_data(row)
        self.empty_rows = self.empty_rows + 1 if row_data.get('columns') is None or row_data == {} else 0
        return row_data

    def __map_row_data(self):
        self.empty_rows = 0
        try:
            rows =  [self.__row_data_wrapper(row) for row in self.current_sheet.iter_rows() if self.empty_rows < 50 and row[0].row < self.current_sheet.max_row+2]
            return [row for row in rows if row.get('columns') is not None]
        except:
            print("Error getting rows")
            return []

    def __map_sheet_data(self, sheet, index):
        try:
            self.current_sheet = sheet
            self.current_sheet_number = index + 1
            self.current_sheet_ranges = self.current_sheet.merged_cells.ranges if self.current_sheet.merged_cells else []

            return {
                "sheetnumber": self.current_sheet_number,
                "sheetname": sheet.title,
                "font": self.__get_default_font_data(),
                "lines": self.__map_row_data()
            }
        except:
            print("Error mapping sheet data", sys.exc_info()[0])
            return {}

    def parse_xlsx_to_json_file(self, excel_path):
        try:
            self.excel_path = excel_path
            self.__open_workbook(excel_path)
            self.__check_for_custom_index(excel_path)

            sheet_data = [self.__map_sheet_data(sheet, index) for index, sheet in enumerate(self.workbook.worksheets)]

            return json.dumps({"sheets": sheet_data}, ensure_ascii=False)
        except Exception as e:
            print(e)
            return json.dumps({"error": str(e)}, ensure_ascii=False)
